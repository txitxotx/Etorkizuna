#!/usr/bin/env python3
"""
actualizar_precios.py
=====================
Actualiza automáticamente el precio de hoy (col H) en cartera_real_gvc.xlsx
para los fondos con ISIN público consultando quefondos.com (sin API key).

USO:
    python actualizar_precios.py
    python actualizar_precios.py mi_cartera.xlsx   # ruta personalizada

FONDOS AUTO-ACTUALIZABLES (tienen ISIN público):
  • Fidelity MSCI World   IE00BYX5NX33  → quefondos.com
  • Fidelity S&P 500      IE00BYX5MX67  → quefondos.com
  • Pictet China Index    LU0625737910  → quefondos.com

FONDOS MANUALES (GVC Gaesco, no tienen feed público):
  → Actualiza tú la col H desde el informe mensual de GVC
"""

import sys, re, time, json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess; subprocess.run([sys.executable,"-m","pip","install","openpyxl","-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

try:
    import urllib.request as urlreq
except: pass

EXCEL_FILE = Path(sys.argv[1]) if len(sys.argv)>1 else Path("cartera_real_gvc.xlsx")

# ── Fondos con ISIN público → precio auto-obtenible ───────────────────────────
AUTO_FUNDS = {
    "IE00BYX5NX33": "Fidelity MSCI World Index P EUR Acc",
    "IE00BYX5MX67": "Fidelity S&P 500 Index P EUR Acc",
    "LU0625737910": "Pictet China Index P EUR",
}

def get_price_quefondos(isin: str) -> tuple[float|None, str]:
    """Obtiene VL de quefondos.com para cualquier fondo con ISIN."""
    url = f"https://www.quefondos.com/es/fondos/ficha/index.html?isin={isin}"
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; portfolio-bot/1.0)",
        "Accept": "text/html",
    }
    try:
        req = urlreq.Request(url, headers=headers)
        with urlreq.urlopen(req, timeout=10) as r:
            html = r.read().decode("utf-8", errors="replace")
        # Pattern: "Valor liquidativo: X,XXXXXX EUR"
        m = re.search(r"Valor liquidativo:\s*([\d,.]+)\s*EUR", html)
        if m:
            price = float(m.group(1).replace(",","."))
            # Date pattern: "Fecha: DD/MM/YYYY"
            d = re.search(r"Fecha:\s*(\d{2}/\d{2}/\d{4})", html)
            date_str = d.group(1) if d else "?"
            return price, date_str
    except Exception as e:
        print(f"    ⚠  Error obteniendo {isin}: {e}")
    return None, "?"

def get_price_finect(isin: str) -> tuple[float|None, str]:
    """Fallback: finect.com"""
    url = f"https://www.finect.com/fondos-inversion/{isin}"
    try:
        req = urlreq.Request(url, headers={"User-Agent":"Mozilla/5.0"})
        with urlreq.urlopen(req, timeout=10) as r:
            html = r.read().decode("utf-8", errors="replace")
        m = re.search(r'"nav"\s*:\s*([\d.]+)', html)
        if m:
            return float(m.group(1)), "finect"
    except: pass
    return None, "?"

def main():
    if not EXCEL_FILE.exists():
        print(f"❌  No se encuentra '{EXCEL_FILE}'")
        sys.exit(1)

    print(f"\n🔄  Actualizando precios en: {EXCEL_FILE}")
    print(f"    Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["📋 ACTIVOS"]

    EUR_FMT = '#,##0.000000\\ "€"'
    updates  = {}   # isin → (price, date)
    found    = {}   # row → isin

    # ── Scan ACTIVOS to find rows that match our ISINs (via col Q notes) ──────
    for r in range(5, 60):
        notas = str(ws.cell(row=r, column=17).value or "")
        for isin in AUTO_FUNDS:
            if isin in notas:
                found[r] = isin

    if not found:
        print("⚠  No se encontraron filas con ISIN en columna Notas (col Q).")
        print("   Asegúrate de tener el Excel correcto (cartera_real_gvc.xlsx).")
        return

    # ── Fetch prices ──────────────────────────────────────────────────────────
    for isin in set(found.values()):
        name = AUTO_FUNDS[isin]
        print(f"  🌐  Consultando {name} ({isin})...", end=" ", flush=True)
        price, date = get_price_quefondos(isin)
        if price is None:
            price, date = get_price_finect(isin)
        if price is not None:
            updates[isin] = (price, date)
            print(f"✅  {price:.6f} € (VL {date})")
        else:
            print("❌  No se pudo obtener precio")
        time.sleep(1)   # cortesía con el servidor

    if not updates:
        print("\n❌  No se pudo obtener ningún precio. Revisa la conexión a internet.")
        return

    # ── Write to Excel ─────────────────────────────────────────────────────────
    updated_rows = []
    for row, isin in found.items():
        if isin not in updates:
            continue
        price, date = updates[isin]
        c = ws.cell(row=row, column=8, value=price)
        c.font = Font(bold=True, size=10, color="0000FF", name="Arial")
        c.fill = PatternFill("solid", fgColor="FFFF99")
        c.number_format = '#,##0.000000\\ "€"'
        c.alignment = Alignment(horizontal="right", vertical="center")

        # Recalculate val, gp, rt for this row
        particip = ws.cell(row=row, column=5).value or 0
        invested = ws.cell(row=row, column=7).value or 0
        if isinstance(particip, (int,float)) and particip:
            new_val = round(particip * price, 2)
            new_gp  = round(new_val - invested, 2)
            new_rt  = new_gp / invested if invested else 0
            ws.cell(row=row, column=9).value = new_val
            ws.cell(row=row, column=10).value = new_gp
            ws.cell(row=row, column=11).value = new_rt
            for col, fmt in [(9,'#,##0.00\\ "€"'),(10,'#,##0.00\\ "€"'),(11,'0.00%')]:
                cc = ws.cell(row=row, column=col)
                cc.number_format = fmt
                cc.font = Font(size=10, color="000000", name="Arial")
                cc.alignment = Alignment(horizontal="right", vertical="center")
            updated_rows.append((row, AUTO_FUNDS[isin], price, date, new_val, new_gp, new_rt))

    # ── Recalculate total row ─────────────────────────────────────────────────
    tot_row = None
    for r in range(5, 60):
        if ws.cell(row=r, column=1).value == "TOTAL CARTERA":
            tot_row = r; break

    if tot_row:
        total_inv = sum(ws.cell(row=r,column=7).value or 0 for r in range(5,tot_row))
        total_val = sum(ws.cell(row=r,column=9).value or 0 for r in range(5,tot_row))
        total_gp  = total_val - total_inv
        total_rt  = total_gp / total_inv if total_inv else 0
        WHITE="E8ECF4"; DARK="1E2430"
        for col,val,fmt in [(7,total_inv,'#,##0.00\\ "€"'),(9,total_val,'#,##0.00\\ "€"'),
                            (10,total_gp,'#,##0.00\\ "€"'),(11,total_rt,'0.00%')]:
            c=ws.cell(row=tot_row,column=col,value=val)
            c.font=Font(bold=True,size=10,color=WHITE,name="Arial")
            c.fill=PatternFill("solid",fgColor=DARK)
            c.number_format=fmt; c.alignment=Alignment(horizontal="right",vertical="center")
        # Update weights
        for r in range(5,tot_row):
            vr=ws.cell(row=r,column=9).value
            if isinstance(vr,(int,float)) and vr and total_val:
                c=ws.cell(row=r,column=14,value=round(vr/total_val,6))
                c.number_format="0.00%"; c.alignment=Alignment(horizontal="right",vertical="center")

    # ── Also update INPUTS sheet metrics ─────────────────────────────────────
    if tot_row:
        wsi = wb["⚙️ INPUTS"]
        import math
        w_rf,w_rv=0.65,0.33; r_rf,r_rv=0.065,0.12; v_rf,v_rv=0.04,0.15; rf=0.02
        exp_ret = w_rf*r_rf + w_rv*r_rv
        exp_vol = math.sqrt((w_rf*v_rf)**2 + (w_rv*v_rv)**2)
        sharpe  = (exp_ret-rf)/exp_vol
        for row,val,fmt in [(18,exp_ret,"0.00%"),(19,exp_vol,"0.00%"),(20,sharpe,"0.00"),
                            (21,total_inv,'#,##0.00\\ "€"'),(22,total_val,'#,##0.00\\ "€"'),
                            (23,total_gp,'#,##0.00\\ "€"'),(24,total_rt,"0.00%"),
                            (25,(1+exp_ret)**10-1,"0.0%")]:
            c=wsi.cell(row=row,column=2,value=val)
            c.font=Font(size=10,color="000000",name="Arial")
            c.number_format=fmt; c.alignment=Alignment(horizontal="center",vertical="center")

    wb.save(EXCEL_FILE)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'─'*60}")
    print(f"✅  ACTUALIZACIÓN COMPLETADA — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'─'*60}")
    for row,name,price,date,val,gp,rt in updated_rows:
        sign = "▲" if rt>=0 else "▼"
        print(f"  {sign} {name[:38]:<38}  {price:>10.4f} €  →  {val:>9.2f} €  ({rt*100:+.2f}%)")
    if tot_row:
        print(f"{'─'*60}")
        print(f"  📊 TOTAL CARTERA:  €{total_val:>10,.2f}  (G/P: €{total_gp:+,.2f} | {total_rt*100:+.2f}%)")
    print(f"\n⚠   Fondos GVC Gaesco: actualiza manualmente la col H desde el informe GVC.")
    print(f"    Luego vuelve a ejecutar este script para recalcular los totales.\n")

if __name__ == "__main__":
    main()
