#!/usr/bin/env python3
"""
parse_excel.py — Lee cartera_real_gvc.xlsx y genera public/data.json
USO: python parse_excel.py [ruta_excel]
"""
import json, sys, os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

EXCEL_FILE  = sys.argv[1] if len(sys.argv) > 1 else "cartera_real_gvc.xlsx"
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "public", "data.json")

SHEET_ASSETS = "📋 ACTIVOS"
SHEET_INPUTS = "⚙️ INPUTS"
SHEET_HIST   = "📈 HISTÓRICO"
SHEET_BYACT  = "📉 HISTÓRICO POR ACTIVO"

def to_float(v, d=0.0):
    if v is None or str(v).strip() in ("","—","#N/A","#REF!","#VALUE!","#DIV/0!","⟵ ACTUALIZAR"):
        return d
    try:
        return float(str(v).replace(",",".").replace("€","").replace("%","").replace(" ","").strip())
    except:
        return d

def to_pct(v, d=0.0):
    f = to_float(v, d)
    return f / 100 if abs(f) > 1.5 else f

def to_str(v): return str(v).strip() if v is not None else ""

def parse():
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: No se encuentra '{EXCEL_FILE}'")
        sys.exit(1)

    print(f"Leyendo {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # ── Activos ────────────────────────────────────────────────────────────────
    ws_act = wb[SHEET_ASSETS]
    assets = []
    for r in range(5, 55):
        row = [ws_act.cell(row=r, column=c).value for c in range(1, 18)]
        name = to_str(row[1])
        cat  = to_str(row[2])
        if not name or name == "TOTAL CARTERA" or cat not in ("RF","RV","SCR"):
            continue
        
        titles   = to_float(row[4])
        buy_px   = to_float(row[5])
        invested = to_float(row[6])
        price_now= row[7]
        curr_val = to_float(row[8])
        gp       = to_float(row[9])
        rt       = to_pct(row[10])
        ytd      = to_pct(row[11])
        mtd      = to_pct(row[12])
        weight   = to_pct(row[13])
        fecha    = to_str(row[14])
        notas    = to_str(row[16])

        # If price is missing, val = invested
        pn = to_float(price_now) if price_now and str(price_now).strip() not in ("⟵ ACTUALIZAR","") else 0
        if curr_val == 0 and invested > 0:
            curr_val = invested
        if gp == 0 and curr_val > 0 and invested > 0:
            gp = curr_val - invested
        if rt == 0 and invested > 0 and gp != 0:
            rt = gp / invested

        assets.append({
            "name": name, "cat": cat,
            "titles": titles, "buy_px": buy_px,
            "invested": invested, "price_now": pn,
            "val": curr_val, "gp": gp,
            "rt": rt, "ytd": ytd, "mtd": mtd, "weight": weight,
            "fecha_inicio": fecha, "notas": notas,
        })

    if not assets:
        print("ERROR: No se encontraron activos.")
        sys.exit(1)

    # ── Inputs ─────────────────────────────────────────────────────────────────
    ws_inp = wb[SHEET_INPUTS]
    def inp_pct(row, col=2): return to_pct(ws_inp.cell(row=row, column=col).value)
    def inp_val(row, col=2): return to_float(ws_inp.cell(row=row, column=col).value)

    inputs = {
        "rf":                 inp_pct(2),
        "market_premium":     inp_pct(3),
        "inflation":          inp_pct(4),
        "tax_rate":           inp_pct(5),
        "fee_rf":             inp_pct(6),
        "fee_rv":             inp_pct(7),
        "target_return":      inp_pct(8),
        "target_vol":         inp_pct(9),
        "target_sharpe":      inp_val(10),
        "target_weight_rf":   inp_pct(13),
        "target_weight_rv":   inp_pct(14),
        "target_weight_scr":  inp_pct(15),
        "exp_ret_rf":         inp_pct(13, 5),
        "exp_ret_rv":         inp_pct(14, 5),
        "exp_vol_rf":         inp_pct(13, 6),
        "exp_vol_rv":         inp_pct(14, 6),
        "sharpe_rf":          inp_val(13, 7),
        "sharpe_rv":          inp_val(14, 7),
    }
    # Portfolio-level (rows 18-25)
    inputs["exp_return_portfolio"] = inp_pct(18)
    inputs["exp_vol_portfolio"]    = inp_pct(19)
    inputs["sharpe_portfolio"]     = inp_val(20)
    inputs["horizon_years"]        = inp_val(23, 5)

    # Fallback calculations if cells are 0 from unrecalculated formulas
    if inputs["exp_return_portfolio"] == 0:
        inputs["exp_return_portfolio"] = (
            inputs["target_weight_rf"] * inputs["exp_ret_rf"] +
            inputs["target_weight_rv"] * inputs["exp_ret_rv"]
        )
    if inputs["exp_vol_portfolio"] == 0:
        import math
        wr, wv = inputs["target_weight_rf"], inputs["target_weight_rv"]
        vr, vv = inputs["exp_vol_rf"], inputs["exp_vol_rv"]
        inputs["exp_vol_portfolio"] = math.sqrt((wr*vr)**2 + (wv*vv)**2) if (vr or vv) else 0.07
    if inputs["sharpe_portfolio"] == 0 and inputs["exp_vol_portfolio"]:
        inputs["sharpe_portfolio"] = (inputs["exp_return_portfolio"] - inputs["rf"]) / inputs["exp_vol_portfolio"]
    if inputs["horizon_years"] == 0:
        inputs["horizon_years"] = 10

    # ── Historical data ────────────────────────────────────────────────────────
    history = []
    try:
        ws_hist = wb[SHEET_HIST]
        for r in range(5, 55):
            fecha = ws_hist.cell(row=r, column=1).value
            val   = to_float(ws_hist.cell(row=r, column=2).value)
            inv   = to_float(ws_hist.cell(row=r, column=3).value)
            if not fecha or val == 0: continue
            gp_h  = to_float(ws_hist.cell(row=r, column=4).value) or (val - inv)
            rt_h  = to_pct(ws_hist.cell(row=r, column=5).value) or (gp_h/inv if inv else 0)
            history.append({
                "date":  str(fecha).strip(),
                "val":   val, "inv": inv, "gp": gp_h, "rt": rt_h,
                "w_rf":  to_pct(ws_hist.cell(row=r, column=6).value),
                "w_rv":  to_pct(ws_hist.cell(row=r, column=7).value),
                "notes": to_str(ws_hist.cell(row=r, column=10).value),
            })
    except: pass

    # ── Per-asset history ──────────────────────────────────────────────────────
    asset_history = {}
    try:
        ws_by = wb[SHEET_BYACT]
        # Row 4 = headers (dates), rows 5+ = assets
        dates = []
        for c in range(2, 15):
            v = ws_by.cell(row=4, column=c).value
            if v: dates.append(to_str(v))
            else: break
        for r in range(5, 40):
            name = to_str(ws_by.cell(row=r, column=1).value)
            if not name: continue
            rents = []
            for ci, dt in enumerate(dates, 2):
                v = ws_by.cell(row=r, column=ci).value
                rents.append({"date": dt, "rt": to_pct(v) if v and str(v).strip() != "—" else None})
            asset_history[name] = rents
    except: pass

    # ── Summary ────────────────────────────────────────────────────────────────
    total_inv = sum(a["invested"] for a in assets)
    total_val = sum(a["val"] for a in assets)
    total_gp  = total_val - total_inv
    total_rt  = total_gp / total_inv if total_inv else 0

    def cat_data(code):
        grp = [a for a in assets if a["cat"] == code]
        inv = sum(a["invested"] for a in grp)
        val = sum(a["val"] for a in grp)
        gp  = val - inv
        ytd = sum(a["ytd"]*a["val"] for a in grp)/val if val else 0
        mtd = sum(a["mtd"]*a["val"] for a in grp)/val if val else 0
        return {"inv":round(inv,2),"val":round(val,2),"gp":round(gp,2),
                "rt":round(gp/inv if inv else 0,6),"ytd":round(ytd,6),"mtd":round(mtd,6),
                "weight":round(val/total_val if total_val else 0,6),"count":len(grp)}

    sorted_rt = sorted(assets, key=lambda a: a["rt"])
    summary = {
        "total_inv":   round(total_inv,2),
        "total_val":   round(total_val,2),
        "total_gp":    round(total_gp,2),
        "total_rt":    round(total_rt,6),
        "updated_at":  datetime.now().strftime("%d/%m/%Y %H:%M"),
        "best_asset":  {"name":sorted_rt[-1]["name"]if assets else "","rt":sorted_rt[-1]["rt"]if assets else 0},
        "worst_asset": {"name":sorted_rt[0]["name"] if assets else "","rt":sorted_rt[0]["rt"] if assets else 0},
        "cats":        {"RF":cat_data("RF"),"RV":cat_data("RV"),"SCR":cat_data("SCR")},
    }

    # Add current snapshot to history if not already present
    today_str = datetime.now().strftime("%d/%m/%Y")
    if not history or history[-1].get("date","") != today_str:
        history.append({
            "date": today_str, "val": total_val, "inv": total_inv,
            "gp": total_gp, "rt": total_rt,
            "w_rf": summary["cats"]["RF"]["weight"],
            "w_rv": summary["cats"]["RV"]["weight"],
            "notes": "Auto-snapshot",
        })

    output = {
        "generated": datetime.now().isoformat(),
        "source":    os.path.basename(EXCEL_FILE),
        "assets":    assets,
        "inputs":    inputs,
        "summary":   summary,
        "history":   history,
        "asset_history": asset_history,
        "scenarios": [],
    }

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    s = summary
    print(f"\n✅  data.json generado: {OUTPUT_FILE}")
    print(f"   Activos:      {len(assets)}")
    print(f"   Total inv:    €{total_inv:>12,.2f}")
    print(f"   Valor actual: €{total_val:>12,.2f}")
    print(f"   G/P total:    €{total_gp:>+12,.2f}  ({total_rt*100:+.2f}%)")
    print(f"   RF:  €{s['cats']['RF']['val']:>10,.0f}  ({s['cats']['RF']['weight']*100:.1f}%)")
    print(f"   RV:  €{s['cats']['RV']['val']:>10,.0f}  ({s['cats']['RV']['weight']*100:.1f}%)")
    print(f"   SCR: €{s['cats']['SCR']['val']:>10,.0f}  ({s['cats']['SCR']['weight']*100:.1f}%)")
    print(f"   Rf:  {inputs['rf']*100:.2f}%  |  Sharpe est: {inputs['sharpe_portfolio']:.2f}")
    print(f"   Histórico: {len(history)} snapshots\n")

if __name__ == "__main__":
    parse()
